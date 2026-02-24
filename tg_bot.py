"""
tg_bot.py — Telegram bot for personal finance logging to Excel (UZS).

FEATURES
- Text input -> parse amount + operation type -> write to Excel
- Supports:
  • "минус 15к такси" -> -15000 UZS
  • "плюс 200к зарплата" -> +200000 UZS
  • "+50000 бонус" / "-7000 кофе"
  • "100 долларов такси" / "$100 такси" / "100 usd такси"
  • "50 евро подарки" / "1000 руб транспорт"
  • Without 'сум' word (defaults to UZS)
- FX conversion via CBU Uzbekistan (daily JSON)
- IMPORTANT: If finance.xlsx is OPEN/LOCKED, bot DOES NOT CRASH.
  It appends rows to pending.csv. When you close Excel, run /sync to flush pending.csv into Excel.

INSTALL
  pip install python-telegram-bot==21.6 openpyxl python-dotenv requests

RUN
  python tg_bot.py

ENV
  Create .env next to this file:
    BOT_TOKEN=123456:ABC...
"""

import os
import re
import csv
import logging
from datetime import datetime
from typing import Optional, Tuple, Dict, Any

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, ContextTypes, filters

# ---------------- CONFIG ----------------
EXCEL_FILE = "finance.xlsx"
SHEET_NAME = "Operations"
PENDING_CSV = "pending.csv"

HEADERS = [
    "date_time",
    "amount_uzs",
    "currency",
    "type",
    "description",
    "orig_amount",
    "orig_currency",
    "rate",
    "raw_text"
]

# Keywords that imply expense/income
EXPENSE_WORDS = [
    "потратил", "потратила", "купил", "купила", "оплатил", "оплатила",
    "расход", "заплатил", "заплатила", "трата"
]
INCOME_WORDS = [
    "пополнение", "пополнил", "пополнила", "получил", "получила",
    "поступление", "приход", "зарплата", "зп", "бонус"
]

# Currency aliases (extend as you like)
CURRENCY_ALIASES = {
    "usd": "USD", "$": "USD", "доллар": "USD", "доллара": "USD", "долларов": "USD",
    "eur": "EUR", "€": "EUR", "евро": "EUR",
    "rub": "RUB", "₽": "RUB", "руб": "RUB", "рубль": "RUB", "рубля": "RUB", "рублей": "RUB",
    "uzs": "UZS", "сум": "UZS", "сўм": "UZS", "som": "UZS", "so'm": "UZS",
}

# FX cache (one fetch per day)
FX_CACHE: Dict[str, Any] = {"date": None, "rates": {}}

# ---------------- LOGGING ----------------
logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO
)
logger = logging.getLogger("finance_bot")

# ---------------- EXCEL ----------------
def _safe_sheet_name(base: str) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", base)
    return base[:31]

def ensure_excel() -> None:
    """
    Ensures Excel file + sheet exists and has correct headers.
    If sheet exists but headers differ, creates a new sheet with timestamp.
    If file is locked/open, it won't crash (it will just return; writes will go to pending.csv).
    """
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
        return

    try:
        wb = load_workbook(EXCEL_FILE)
    except PermissionError:
        # Excel is open/locked. We'll still allow bot to run; it will queue to pending.csv.
        logger.warning("'%s' is locked/open. Bot will queue writes to %s.", EXCEL_FILE, PENDING_CSV)
        return

    if SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADERS)
        try:
            wb.save(EXCEL_FILE)
        except PermissionError:
            logger.warning("'%s' became locked/open while saving. Will queue writes.", EXCEL_FILE)
        return

    ws = wb[SHEET_NAME]
    existing = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if existing != HEADERS:
        ts = datetime.now().strftime("%H%M%S")
        new_name = _safe_sheet_name(f"{SHEET_NAME}_{ts}")
        ws2 = wb.create_sheet(new_name)
        ws2.append(HEADERS)
        logger.warning(
            "Headers mismatch in sheet '%s'. Created new sheet '%s' with correct headers.",
            SHEET_NAME, new_name
        )
        try:
            wb.save(EXCEL_FILE)
        except PermissionError:
            logger.warning("'%s' locked/open; cannot save header-fix now. Will queue writes.", EXCEL_FILE)

def append_pending_row(row: list) -> None:
    file_exists = os.path.exists(PENDING_CSV)
    with open(PENDING_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if not file_exists:
            w.writerow(HEADERS)
        w.writerow(row)

def flush_pending_to_excel() -> tuple[int, str]:
    """
    Flush pending.csv into Excel file.
    Returns (rows_written, message).
    """
    if not os.path.exists(PENDING_CSV):
        return 0, "Очередь пустая: pending.csv не найден."

    try:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        ws = wb[SHEET_NAME]
    except FileNotFoundError:
        ensure_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
    except PermissionError:
        return 0, f"Excel-файл '{EXCEL_FILE}' открыт. Закрой его и повтори /sync."

    rows_written = 0
    with open(PENDING_CSV, "r", newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        _ = next(r, None)  # skip headers
        for row in r:
            ws.append(row)
            rows_written += 1

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        return 0, f"Excel-файл '{EXCEL_FILE}' открыт. Закрой его и повтори /sync."

    os.remove(PENDING_CSV)
    return rows_written, f"✅ Синхронизировал {rows_written} строк(и) в {EXCEL_FILE}."

def append_row(
    dt: str,
    amount_uzs: int,
    currency: str,
    op_type: str,
    desc: str,
    orig_amount: float,
    orig_currency: str,
    rate: float,
    raw: str,
) -> str:
    """
    Tries to write to Excel. If Excel is locked/open -> appends to pending.csv (no crash).
    Returns: "excel" or "pending"
    """
    row = [
    dt,
    amount_uzs,
    currency,
    op_type,
    desc,
    orig_amount,
    orig_currency,
    rate,
    raw
]
    try:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        ws = wb[SHEET_NAME]
        ws.append(row)
        wb.save(EXCEL_FILE)
        return "excel"
    except PermissionError:
        append_pending_row(row)
        return "pending"
    except FileNotFoundError:
        ensure_excel()
        # try again once
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            ws.append(row)
            wb.save(EXCEL_FILE)
            return "excel"
        except Exception:
            append_pending_row(row)
            return "pending"

# ---------------- FX (CBU) ----------------
def get_cbu_rates() -> dict:
    """
    Returns rates dict: {"USD": 12169.41, "EUR": ..., "RUB": ...}
    Cached per day.
    """
    today = datetime.now().strftime("%Y-%m-%d")
    if FX_CACHE["date"] == today and FX_CACHE["rates"]:
        return FX_CACHE["rates"]

    url = "https://cbu.uz/ru/arkhiv-kursov-valyut/json/"
    resp = requests.get(url, timeout=10)
    resp.raise_for_status()
    data = resp.json()

    rates = {}
    for item in data:
        ccy = item.get("Ccy")
        rate_str = item.get("Rate")
        if ccy and rate_str:
            try:
                rates[str(ccy).upper()] = float(str(rate_str).replace(",", "."))
            except ValueError:
                continue

    FX_CACHE["date"] = today
    FX_CACHE["rates"] = rates
    return rates

def convert_to_uzs(amount: float, currency: str) -> Tuple[float, float]:
    """
    amount in currency -> returns (amount_uzs, rate)
    where rate = UZS per 1 unit of currency
    """
    currency = currency.upper()
    if currency == "UZS":
        return amount, 1.0

    rates = get_cbu_rates()
    if currency not in rates:
        raise ValueError(f"No rate for {currency}")

    rate = float(rates[currency])
    return amount * rate, rate

# ---------------- PARSING ----------------
def detect_currency(t: str) -> str:
    # Prioritize symbol matches
    for sym in ["$", "€", "₽"]:
        if sym in t:
            return CURRENCY_ALIASES[sym]
    # Word matches
    for k, v in CURRENCY_ALIASES.items():
        if k in ["$", "€", "₽"]:
            continue
        if re.search(rf"\b{re.escape(k)}\b", t):
            return v
    return "UZS"

def _detect_op_type(t: str) -> str:
    s = t.strip()
    if s.startswith("+"):
        return "income"
    if s.startswith("-"):
        return "expense"
    if "плюс" in t:
        return "income"
    if "минус" in t:
        return "expense"
    if any(w in t for w in INCOME_WORDS):
        return "income"
    if any(w in t for w in EXPENSE_WORDS):
        return "expense"
    return "expense"

def parse_text(text: str) -> Optional[Tuple[int, str, str, str, float, str, float]]:
    """
    Returns:
    (amount_uzs_signed_int, "UZS", op_type, description, orig_amount_float, orig_currency, rate)
    """
    t = text.lower().strip()
    op_type = _detect_op_type(t)
    currency = detect_currency(t)

    # Amount (first number). Supports 'к' suffix = *1000
    m = re.search(r"([+-]?\d[\d\s.,]*)(к)?", t)
    if not m:
        return None

    num_raw = m.group(1).strip()
    has_k = bool(m.group(2))

    # number sign overrides op_type
    if num_raw.startswith("+"):
        op_type = "income"
    elif num_raw.startswith("-"):
        op_type = "expense"

    cleaned = re.sub(r"[^\d.,]", "", num_raw).replace(",", ".")
    if cleaned.count(".") > 1:
        parts = cleaned.split(".")
        cleaned = "".join(parts[:-1]) + "." + parts[-1]

    try:
        orig_amount = float(cleaned)
    except ValueError:
        return None

    if has_k:
        orig_amount *= 1000

    # Convert to UZS
    try:
        amount_uzs, rate = convert_to_uzs(orig_amount, currency)
    except Exception as e:
        logger.warning("FX conversion failed (%s). Treating as UZS.", str(e))
        amount_uzs, rate, currency = orig_amount, 1.0, "UZS"

    signed_uzs = amount_uzs if op_type == "income" else -amount_uzs

    # Description after amount
    desc_part = t[m.end():].strip()
    desc_part = re.sub(r"^(на|за|для)\s+", "", desc_part)
    description = desc_part if desc_part else "без описания"

    return int(round(signed_uzs)), "UZS", op_type, description, orig_amount, currency, float(rate)

# ---------------- TELEGRAM HANDLERS ----------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Пиши расходы/доходы, например:\n"
        "• минус 15к такси\n"
        "• потратил 15000 на еду\n"
        "• плюс 200к зарплата\n"
        "• 100 долларов такси / $100 такси\n\n"
        "Если Excel-файл открыт, я записываю в очередь. Потом закрой Excel и введи /sync."
    )

async def sync(update: Update, context: ContextTypes.DEFAULT_TYPE):
    n, msg = flush_pending_to_excel()
    await update.message.reply_text(msg)

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text or ""
    parsed = parse_text(text)

    if not parsed:
        await update.message.reply_text(
            "Не понял сумму 😕\n"
            "Пример: 'минус 15к такси' или '100 долларов такси' или '+50000 бонус'"
        )
        return

    amount_uzs, currency_out, op_type, desc, orig_amount, orig_currency, rate = parsed
    dt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    status = append_row(dt, amount_uzs, currency_out, op_type, desc, orig_amount, orig_currency, rate, text)

    prefix = "✅ Записал" if status == "excel" else "🕒 Excel открыт — записал в очередь"
    sign = "+" if amount_uzs > 0 else ""

    if orig_currency != "UZS":
        await update.message.reply_text(
            f"{prefix}: {sign}{amount_uzs} UZS — {desc}\n"
            f"({orig_amount:g} {orig_currency} × {rate:g})"
        )
    else:
        await update.message.reply_text(f"{prefix}: {sign}{amount_uzs} UZS — {desc}")

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Exception while handling an update: %s", context.error)

# ---------------- MAIN ----------------
# ---------------- MAIN ----------------
def main():
    load_dotenv()
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise RuntimeError("Не найден BOT_TOKEN. Добавь BOT_TOKEN в .env или переменные окружения.")

    ensure_excel()

    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("sync", sync))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_error_handler(error_handler)

    print("Bot is running...")
    
    webhook_url = os.environ.get("WEBHOOK_URL")

    if webhook_url:
        port = int(os.environ.get("PORT", 8000))
        app.run_webhook(
            listen="0.0.0.0",
            port=port,
            webhook_url=webhook_url,
        )
    else:
        app.run_polling()

if __name__ == "__main__":
    main()